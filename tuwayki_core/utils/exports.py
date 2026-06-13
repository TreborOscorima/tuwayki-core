"""Utilidades de exportacion a Excel y PDF."""
import io
import datetime
import logging
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from typing import Any
from decimal import Decimal, InvalidOperation

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
THICK_BORDER_BOTTOM = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="medium"),
)

TITLE_FONT = Font(bold=True, size=14, color="4F46E5")
SUBTITLE_FONT = Font(bold=True, size=11, color="374151")
NOTE_FONT = Font(italic=True, size=9, color="6B7280")
TOTAL_FILL = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
POSITIVE_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
NEGATIVE_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

CURRENCY_FORMAT = '"S/"#,##0.00'
PERCENT_FORMAT = '0.00%'
NUMBER_FORMAT = '#,##0'
PDF_CONTROL_KEYS = {"logo_path", "logo", "page_size", "column_widths", "wrap_columns"}


def _sanitize_excel_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    stripped = value.lstrip()
    if stripped.startswith(("=", "+", "-", "@")):
        return f"'{value}"
    return value


def _safe_decimal(value: Any) -> Decimal:
    if value is None:
        return Decimal("0")
    try:
        return Decimal(str(value))
    except (ValueError, TypeError, InvalidOperation):
        return Decimal("0")


def _safe_float(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def add_company_header(
    ws: Worksheet,
    company_name: str,
    report_title: str,
    period_str: str = "",
    columns: int = 6,
    generated_at: datetime.datetime | None = None,
) -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    title_cell = ws.cell(row=1, column=1, value=company_name.upper() if company_name else "EMPRESA")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=columns)
    subtitle_cell = ws.cell(row=2, column=1, value=report_title)
    subtitle_cell.font = SUBTITLE_FONT
    subtitle_cell.alignment = Alignment(horizontal="center")

    if period_str:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=columns)
        period_cell = ws.cell(row=3, column=1, value=f"📅 {period_str}")
        period_cell.font = Font(size=10, color="6B7280")
        period_cell.alignment = Alignment(horizontal="center")

    now = generated_at or datetime.datetime.now()
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=columns)
    date_cell = ws.cell(row=4, column=1, value=f"⏱ Generado: {now.strftime('%d/%m/%Y a las %H:%M:%S')}")
    date_cell.font = Font(size=9, italic=True, color="9CA3AF")
    date_cell.alignment = Alignment(horizontal="center")

    ws.row_dimensions[5].height = 10
    return 6


def add_totals_row_with_formulas(
    ws: Worksheet,
    row: int,
    start_data_row: int,
    columns_config: list[dict]
) -> None:
    for col_idx, config in enumerate(columns_config, start=1):
        cell = ws.cell(row=row, column=col_idx)
        col_type = config.get("type", "text")

        if col_type == "label":
            cell.value = config.get("value", "TOTAL")
        elif col_type == "sum":
            col_letter = config.get("col_letter", get_column_letter(col_idx))
            cell.value = f"=SUM({col_letter}{start_data_row}:{col_letter}{row-1})"
        elif col_type == "count":
            col_letter = config.get("col_letter", get_column_letter(col_idx))
            cell.value = f"=COUNT({col_letter}{start_data_row}:{col_letter}{row-1})"
        elif col_type == "average":
            col_letter = config.get("col_letter", get_column_letter(col_idx))
            cell.value = f"=AVERAGE({col_letter}{start_data_row}:{col_letter}{row-1})"
        elif col_type == "formula":
            cell.value = config.get("value", "")
        elif col_type == "text":
            cell.value = config.get("value", "")

        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THICK_BORDER_BOTTOM

        if "number_format" in config:
            cell.number_format = config["number_format"]


def add_notes_section(
    ws: Worksheet,
    after_row: int,
    notes: list[str],
    columns: int = 6
) -> int:
    row = after_row + 2

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
    notes_title = ws.cell(row=row, column=1, value="📋 NOTAS Y DEFINICIONES:")
    notes_title.font = Font(bold=True, size=10, color="374151")
    row += 1

    for note in notes:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=columns)
        note_cell = ws.cell(row=row, column=1, value=f"• {note}")
        note_cell.font = NOTE_FONT
        row += 1

    return row


def create_excel_workbook(title: str) -> tuple[Workbook, Worksheet]:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title[:31]
    return workbook, sheet


def style_header_row(ws: Worksheet, row: int, columns: list[str]) -> None:
    for col_idx, header in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def add_data_rows(
    ws: Worksheet,
    data: list[list[Any]],
    start_row: int,
    apply_border: bool = True,
) -> int:
    current_row = start_row
    for row_data in data:
        for col_idx, value in enumerate(row_data, start=1):
            safe_value = _sanitize_excel_value(value)
            cell = ws.cell(row=current_row, column=col_idx, value=safe_value)
            if apply_border:
                cell.border = THIN_BORDER
        current_row += 1
    return current_row


def auto_adjust_column_widths(ws: Worksheet, min_width: int = 10, max_width: int = 50) -> None:
    from openpyxl.cell.cell import MergedCell

    for column in ws.columns:
        max_length = 0
        column_letter = None

        for cell in column:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                break

        if column_letter is None:
            continue

        for cell in column:
            if isinstance(cell, MergedCell):
                continue
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except (ValueError, TypeError, AttributeError):
                pass

        adjusted_width = (max_length + 2)
        if adjusted_width < min_width:
            adjusted_width = min_width
        if adjusted_width > max_width:
            adjusted_width = max_width
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_wrap_text(
    ws: Worksheet,
    columns: list[int],
    start_row: int,
    end_row: int | None = None,
    vertical: str = "top",
) -> None:
    if end_row is None:
        end_row = ws.max_row
    for row in range(start_row, end_row + 1):
        for col in columns:
            cell = ws.cell(row=row, column=col)
            current = cell.alignment or Alignment()
            cell.alignment = current.copy(wrap_text=True, vertical=vertical)


def create_pdf_report(
    buffer: io.BytesIO,
    title: str,
    data: list[list[Any]],
    headers: list[str],
    info_dict: dict[str, Any] | None = None,
) -> None:
    info_dict = info_dict or {}
    page_size = A4 if A4 else letter
    if info_dict.get("page_size") == "letter":
        page_size = letter

    logo_path = info_dict.get("logo_path") or info_dict.get("logo")

    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        leftMargin=36,
        rightMargin=36,
        topMargin=36,
        bottomMargin=36,
    )
    styles = getSampleStyleSheet()
    title_style = styles["Heading1"]
    info_style = ParagraphStyle(
        name="InfoStyle",
        parent=styles["Normal"],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor("#374151"),
    )

    elements: list[Any] = []

    if logo_path:
        try:
            logo = Image(logo_path)
            logo.drawHeight = 50
            logo.drawWidth = 50
            header_table = Table(
                [[logo, Paragraph(title, title_style)]],
                colWidths=[60, doc.width - 60],
            )
            header_table.setStyle(
                TableStyle(
                    [
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 0),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            elements.append(header_table)
        except Exception:
            logging.getLogger(__name__).warning("PDF header table failed, using plain text", exc_info=True)
            elements.append(Paragraph(title, title_style))
    else:
        elements.append(Paragraph(title, title_style))

    elements.append(Spacer(1, 12))

    info_rows: list[list[Any]] = []
    for key, value in info_dict.items():
        if key in PDF_CONTROL_KEYS:
            continue
        label = str(key).replace("_", " ").title()
        info_rows.append(
            [
                Paragraph(f"<b>{label}</b>", info_style),
                Paragraph(str(value), info_style),
            ]
        )

    if info_rows:
        info_table = Table(
            info_rows,
            colWidths=[140, doc.width - 140],
            hAlign="LEFT",
        )
        info_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, -1), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ]
            )
        )
        elements.append(info_table)
        elements.append(Spacer(1, 12))

    table_data: list[list[Any]] = []
    if headers:
        table_data.append(headers)

    column_count = len(headers) if headers else max((len(row) for row in data), default=0)
    wrap_columns = info_dict.get("wrap_columns") or []
    if isinstance(wrap_columns, tuple):
        wrap_columns = list(wrap_columns)
    if not isinstance(wrap_columns, list):
        wrap_columns = []
    wrap_columns = [int(col) for col in wrap_columns if isinstance(col, (int, float))]
    wrap_style = ParagraphStyle(
        name="WrapStyle",
        parent=styles["Normal"],
        fontSize=8,
        leading=11,
        textColor=colors.black,
        wordWrap="CJK",
    )

    for row in data:
        normalized_row = list(row)
        if column_count:
            if len(normalized_row) < column_count:
                normalized_row += [""] * (column_count - len(normalized_row))
            elif len(normalized_row) > column_count:
                normalized_row = normalized_row[:column_count]
        formatted_row: list[Any] = []
        for idx, value in enumerate(normalized_row):
            clean_value = "" if value is None else value
            if idx in wrap_columns and not isinstance(clean_value, Paragraph):
                formatted_row.append(Paragraph(str(clean_value), wrap_style))
            else:
                formatted_row.append(clean_value)
        table_data.append(formatted_row)

    if not table_data:
        table_data = [["Sin datos"]]

    col_widths = None
    column_widths = info_dict.get("column_widths")
    if (
        column_widths
        and column_count
        and isinstance(column_widths, (list, tuple))
        and len(column_widths) == column_count
    ):
        try:
            numeric_widths = [float(width) for width in column_widths]
            total_width = sum(numeric_widths)
            if total_width <= 1.1:
                col_widths = [doc.width * width for width in numeric_widths]
            else:
                col_widths = numeric_widths
        except (TypeError, ValueError):
            col_widths = None

    main_table = Table(
        table_data,
        colWidths=col_widths,
        hAlign="LEFT",
        repeatRows=1 if headers else 0,
    )
    table_style = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]
    )

    for row_idx in range(1, len(table_data)):
        if row_idx % 2 == 0:
            table_style.add(
                "BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#F3F4F6")
            )

    main_table.setStyle(table_style)
    elements.append(main_table)

    doc.build(elements)
    buffer.seek(0)
